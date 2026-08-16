from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook

from aplicacion.base_datos.conexion import SessionLocal
from aplicacion.maestros.categorias.modelos import Categoria
from aplicacion.maestros.marcas.modelos import Marca
from aplicacion.maestros.productos.modelos import Producto
from aplicacion.maestros.productos.servicios import ServicioProducto
from aplicacion.maestros.unidades_medida.repositorio import (
    UnidadMedidaRepositorio,
)

COLUMNAS = (
    ("codigo", "Código"),
    ("nombre", "Nombre"),
    ("codigo_barras", "Código de barras"),
    ("categoria", "Categoría (nombre exacto)"),
    ("marca", "Marca (nombre exacto)"),
    ("unidad_medida", "Unidad de medida (código)"),
    ("precio_venta", "Precio de venta"),
    ("costo", "Costo"),
    ("stock_minimo", "Stock mínimo"),
    ("precio_incluye_iva", "Precio incluye IVA (Sí/No)"),
)

_CAMPOS_NUMERICOS = {
    "precio_venta",
    "costo",
    "stock_minimo",
}

_FILA_EJEMPLO = (
    "PRD-0001",
    "Producto de ejemplo",
    "",
    "",
    "",
    "",
    50000,
    30000,
    0,
    "No",
)


@dataclass
class ResultadoImportacion:

    creados: int = 0
    actualizados: int = 0
    errores: list[tuple[int, str]] = field(
        default_factory=list,
    )

    @property
    def total_ok(self) -> int:

        return self.creados + self.actualizados


def generar_plantilla(destino: Path) -> None:

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Productos"

    hoja.append(
        [encabezado for _, encabezado in COLUMNAS],
    )

    hoja.append(
        list(_FILA_EJEMPLO),
    )

    from openpyxl.utils import get_column_letter

    for indice, (_, encabezado) in enumerate(
        COLUMNAS,
        start=1,
    ):

        hoja.column_dimensions[
            get_column_letter(indice)
        ].width = max(
            len(encabezado),
            18,
        )

    libro.save(destino)


def _valor_celda(valor, campo: str):

    if valor is None:

        return 0 if campo in _CAMPOS_NUMERICOS else ""

    if campo in _CAMPOS_NUMERICOS:

        try:

            return float(valor)

        except (TypeError, ValueError):

            return 0

    return str(valor).strip()


def _fila_a_datos(fila: tuple) -> dict:

    datos = {}

    for indice, (campo, _) in enumerate(COLUMNAS):

        valor = fila[indice] if indice < len(fila) else None

        datos[campo] = _valor_celda(valor, campo)

    return datos


def _fila_vacia(fila: tuple) -> bool:

    return not any(
        str(valor).strip()
        for valor in fila
        if valor is not None
    )


def _resolver_categoria_id(nombre: str) -> int | None:

    if not nombre:

        return None

    db = SessionLocal()

    try:

        categoria = (
            db.query(Categoria)
            .filter(Categoria.nombre.ilike(nombre))
            .first()
        )

    finally:

        db.close()

    if categoria is None:

        raise ValueError(
            f"No existe la categoría '{nombre}'.",
        )

    return categoria.id


def _resolver_marca_id(nombre: str) -> int | None:

    if not nombre:

        return None

    db = SessionLocal()

    try:

        marca = (
            db.query(Marca)
            .filter(Marca.nombre.ilike(nombre))
            .first()
        )

    finally:

        db.close()

    if marca is None:

        raise ValueError(
            f"No existe la marca '{nombre}'.",
        )

    return marca.id


def _resolver_unidad_medida_id(codigo: str) -> int | None:

    if not codigo:

        return None

    unidad = UnidadMedidaRepositorio.obtener_por_codigo(
        codigo,
    )

    if unidad is None:

        raise ValueError(
            f"No existe la unidad de medida '{codigo}'.",
        )

    return unidad.id


def _preparar_datos_producto(datos: dict) -> dict:

    preparado = {
        "codigo": datos["codigo"],
        "nombre": datos["nombre"],
        "codigo_barras": datos["codigo_barras"] or None,
        "precio_venta": datos["precio_venta"],
        "costo": datos["costo"],
        "stock_minimo": datos["stock_minimo"],
        "precio_incluye_iva": str(
            datos.get("precio_incluye_iva") or "",
        ).strip().lower()
        in ("si", "sí", "s", "true", "1"),
    }

    categoria_id = _resolver_categoria_id(
        datos.get("categoria", ""),
    )

    if categoria_id is not None:

        preparado["categoria_id"] = categoria_id

    marca_id = _resolver_marca_id(
        datos.get("marca", ""),
    )

    if marca_id is not None:

        preparado["marca_id"] = marca_id

    unidad_medida_id = _resolver_unidad_medida_id(
        datos.get("unidad_medida", ""),
    )

    if unidad_medida_id is not None:

        preparado["unidad_medida_id"] = unidad_medida_id

    return preparado


def importar_desde_excel(ruta: Path) -> ResultadoImportacion:

    libro = load_workbook(ruta, data_only=True)

    hoja = libro.active

    resultado = ResultadoImportacion()

    for indice_fila, fila in enumerate(
        hoja.iter_rows(min_row=2, values_only=True),
        start=2,
    ):

        if _fila_vacia(fila):

            continue

        datos = _fila_a_datos(fila)

        try:

            producto_datos = _preparar_datos_producto(datos)

            db = SessionLocal()

            try:

                existente = (
                    db.query(Producto)
                    .filter(
                        Producto.codigo.ilike(
                            producto_datos["codigo"],
                        ),
                    )
                    .first()
                )

                existente_id = (
                    existente.id
                    if existente is not None
                    else None
                )

            finally:

                db.close()

            if existente_id is not None:

                ServicioProducto.actualizar(
                    existente_id,
                    producto_datos,
                )

                resultado.actualizados += 1

            else:

                ServicioProducto.guardar(producto_datos)

                resultado.creados += 1

        except ValueError as error:

            resultado.errores.append(
                (indice_fila, str(error)),
            )

        except Exception as error:  # noqa: BLE001

            resultado.errores.append(
                (indice_fila, str(error)),
            )

    return resultado
