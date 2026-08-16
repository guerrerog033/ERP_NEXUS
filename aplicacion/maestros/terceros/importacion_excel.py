from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook

from aplicacion.maestros.terceros.repositorio import (
    TerceroRepositorio,
)
from aplicacion.maestros.terceros.servicio import (
    TerceroServicio,
)

COLUMNAS = (
    ("tipo_documento", "Tipo Documento (CC/CE/NIT/TI/PAS)"),
    ("numero_documento", "Número Documento"),
    ("tipo_tercero", "Tipo de Tercero (Cliente/Proveedor/Otro)"),
    ("razon_social", "Razón Social / Nombre Comercial"),
    ("primer_nombre", "Primer Nombre"),
    ("segundo_nombre", "Segundo Nombre"),
    ("primer_apellido", "Primer Apellido"),
    ("segundo_apellido", "Segundo Apellido"),
    ("direccion", "Dirección"),
    ("ciudad", "Ciudad"),
    ("departamento", "Departamento"),
    ("pais", "País"),
    ("telefono", "Teléfono"),
    ("celular", "Celular"),
    ("correo", "Correo"),
    ("dias_credito", "Días de crédito"),
    ("cupo_credito", "Cupo de crédito"),
)

_CAMPOS_NUMERICOS = {
    "dias_credito",
    "cupo_credito",
}

_FILA_EJEMPLO = (
    "NIT",
    "900123456",
    "Cliente",
    "Cliente de ejemplo S.A.S.",
    "",
    "",
    "",
    "",
    "Calle 1 # 2-3",
    "Bogotá",
    "Cundinamarca",
    "Colombia",
    "",
    "3001234567",
    "cliente@ejemplo.com",
    30,
    5000000,
)


@dataclass
class ResultadoImportacion:

    creados: int = 0
    actualizados: int = 0
    errores: list[tuple[int, str]] = field(
        default_factory=list,
    )

    @property
    def total_ok(self) -> int:

        return self.creados + self.actualizados


def generar_plantilla(destino: Path) -> None:

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Terceros"

    hoja.append(
        [encabezado for _, encabezado in COLUMNAS],
    )

    hoja.append(
        list(_FILA_EJEMPLO),
    )

    from openpyxl.utils import get_column_letter

    for indice, (_, encabezado) in enumerate(
        COLUMNAS,
        start=1,
    ):

        hoja.column_dimensions[
            get_column_letter(indice)
        ].width = max(
            len(encabezado),
            18,
        )

    libro.save(destino)


def _valor_celda(
    valor,
    campo: str,
):

    if valor is None:

        return 0 if campo in _CAMPOS_NUMERICOS else ""

    if campo in _CAMPOS_NUMERICOS:

        try:

            return float(valor)

        except (TypeError, ValueError):

            return 0

    return str(valor).strip()


def _fila_a_datos(fila: tuple) -> dict:

    datos = {}

    for indice, (campo, _) in enumerate(COLUMNAS):

        valor = fila[indice] if indice < len(fila) else None

        datos[campo] = _valor_celda(valor, campo)

    return datos


def _fila_vacia(fila: tuple) -> bool:

    return not any(
        str(valor).strip()
        for valor in fila
        if valor is not None
    )


def importar_desde_excel(
    ruta: Path,
) -> ResultadoImportacion:

    libro = load_workbook(
        ruta,
        data_only=True,
    )

    hoja = libro.active

    resultado = ResultadoImportacion()

    for indice_fila, fila in enumerate(
        hoja.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):

        if _fila_vacia(fila):

            continue

        datos = _fila_a_datos(fila)

        try:

            existente = (
                TerceroRepositorio.obtener_por_documento(
                    datos.get("tipo_documento", ""),
                    datos.get("numero_documento", ""),
                )
                if datos.get("numero_documento")
                else None
            )

            if existente is not None:

                TerceroServicio.actualizar(
                    existente.id,
                    datos,
                )

                resultado.actualizados += 1

            else:

                TerceroServicio.guardar(datos)

                resultado.creados += 1

        except ValueError as error:

            resultado.errores.append(
                (indice_fila, str(error)),
            )

        except Exception as error:  # noqa: BLE001

            resultado.errores.append(
                (indice_fila, str(error)),
            )

    return resultado
