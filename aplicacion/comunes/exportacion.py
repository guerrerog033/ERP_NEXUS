from __future__ import annotations

from pathlib import Path

from PySide6.QtCore import QSizeF
from PySide6.QtGui import (
    QPageSize,
    QTextDocument,
)
from PySide6.QtPrintSupport import QPrinter
from PySide6.QtWidgets import (
    QFileDialog,
    QMessageBox,
    QTableWidget,
    QWidget,
)

try:

    from openpyxl import Workbook

    _TIENE_OPENPYXL = True

except ImportError:

    _TIENE_OPENPYXL = False


def exportar_tabla_excel(
    tabla: QTableWidget,
    *,
    parent: QWidget | None,
    titulo: str,
    nombre_archivo: str | None = None,
) -> bool:

    if tabla.rowCount() == 0:

        QMessageBox.warning(
            parent,
            "Exportar Excel",
            "No hay datos para exportar.",
        )

        return False

    sugerido = (
        nombre_archivo
        or titulo.replace(" ", "_").lower()
    )

    if _TIENE_OPENPYXL:

        filtro = "Excel (*.xlsx)"
        extension = ".xlsx"

    else:

        filtro = "CSV para Excel (*.csv)"
        extension = ".csv"

    if not sugerido.endswith(
        extension,
    ):

        sugerido = f"{sugerido}{extension}"

    ruta, _ = QFileDialog.getSaveFileName(
        parent,
        "Exportar a Excel",
        sugerido,
        filtro,
    )

    if not ruta:

        return False

    destino = Path(ruta)

    if destino.suffix.lower() not in (
        ".xlsx",
        ".csv",
    ):

        destino = destino.with_suffix(
            extension,
        )

    try:

        if destino.suffix.lower() == ".xlsx":

            _guardar_xlsx(
                tabla,
                destino,
            )

        else:

            _guardar_csv(
                tabla,
                destino,
            )

    except OSError as error:

        QMessageBox.critical(
            parent,
            "Exportar Excel",
            f"No se pudo guardar el archivo:\n{error}",
        )

        return False

    QMessageBox.information(
        parent,
        "Exportar Excel",
        f"Archivo exportado:\n{destino}",
    )

    return True


def _encabezados(
    tabla: QTableWidget,
) -> list[str]:

    columnas = []

    for indice in range(
        tabla.columnCount(),
    ):

        item = tabla.horizontalHeaderItem(
            indice,
        )

        columnas.append(
            item.text()
            if item is not None
            else "",
        )

    return columnas


def _filas_texto(
    tabla: QTableWidget,
) -> list[list[str]]:

    filas: list[list[str]] = []

    for fila in range(
        tabla.rowCount(),
    ):

        valores = []

        for columna in range(
            tabla.columnCount(),
        ):

            item = tabla.item(
                fila,
                columna,
            )

            valores.append(
                item.text()
                if item is not None
                else "",
            )

        filas.append(
            valores,
        )

    return filas


def _guardar_xlsx(
    tabla: QTableWidget,
    destino: Path,
) -> None:

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Datos"

    encabezados = _encabezados(
        tabla,
    )

    hoja.append(
        encabezados,
    )

    for fila in _filas_texto(
        tabla,
    ):

        hoja.append(
            fila,
        )

    for indice, encabezado in enumerate(
        encabezados,
        start=1,
    ):

        from openpyxl.utils import (
            get_column_letter,
        )

        columna = get_column_letter(
            indice,
        )

        hoja.column_dimensions[
            columna
        ].width = max(
            len(
                encabezado,
            ),
            12,
        )

    libro.save(
        destino,
    )


def _guardar_csv(
    tabla: QTableWidget,
    destino: Path,
) -> None:

    import csv

    with destino.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as archivo:

        escritor = csv.writer(
            archivo,
            delimiter=";",
        )

        escritor.writerow(
            _encabezados(
                tabla,
            ),
        )

        escritor.writerows(
            _filas_texto(
                tabla,
            ),
        )


def boton_exportar_excel(
    tabla: QTableWidget,
    *,
    parent: QWidget | None,
    titulo: str,
):

    from PySide6.QtWidgets import QPushButton

    boton = QPushButton(
        "Exportar Excel",
    )

    boton.clicked.connect(

        lambda: exportar_tabla_excel(
            tabla,
            parent=parent,
            titulo=titulo,
        ),

    )

    return boton


def _valor_registro(
    registro,
    nombre: str,
):

    if isinstance(
        registro,
        dict,
    ):

        return registro.get(
            nombre,
        )

    return getattr(
        registro,
        nombre,
        None,
    )


def _filas_desde_registros(
    definition,
    registros: list,
) -> tuple[
    list[str],
    list[list[str]],
]:

    from aplicacion.framework.table.column_registry import (
        ColumnRegistry,
    )

    columnas = [
        columna
        for columna in definition.obtener_columnas()
        if columna.visible
    ]

    encabezados = [
        columna.encabezado
        for columna in columnas
    ]

    filas: list[list[str]] = []

    for registro in registros:

        fila = []

        for columna in columnas:

            valor = _valor_registro(
                registro,
                columna.nombre,
            )

            fila.append(
                ColumnRegistry.formatear_valor(
                    columna.widget,
                    valor,
                    columna,
                ),
            )

        filas.append(
            fila,
        )

    return encabezados, filas


def _guardar_xlsx_datos(
    encabezados: list[str],
    filas: list[list[str]],
    destino: Path,
) -> None:

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Datos"

    hoja.append(
        encabezados,
    )

    for fila in filas:

        hoja.append(
            fila,
        )

    from openpyxl.utils import (
        get_column_letter,
    )

    for indice, encabezado in enumerate(
        encabezados,
        start=1,
    ):

        columna = get_column_letter(
            indice,
        )

        hoja.column_dimensions[
            columna
        ].width = max(
            len(
                encabezado,
            ),
            12,
        )

    libro.save(
        destino,
    )


def exportar_registros(
    definition,
    registros: list,
    *,
    parent: QWidget | None,
    titulo: str,
    nombre_archivo: str | None = None,
) -> bool:

    if not registros:

        QMessageBox.warning(
            parent,
            "Exportar Excel",
            "No hay datos para exportar.",
        )

        return False

    sugerido = (
        nombre_archivo
        or titulo.replace(" ", "_").lower()
    )

    if _TIENE_OPENPYXL:

        filtro = "Excel (*.xlsx);;CSV (*.csv)"
        extension = ".xlsx"

    else:

        filtro = "CSV (*.csv)"
        extension = ".csv"

    if not sugerido.endswith(
        (
            ".xlsx",
            ".csv",
        ),
    ):

        sugerido = f"{sugerido}{extension}"

    ruta, _ = QFileDialog.getSaveFileName(
        parent,
        "Exportar a Excel",
        sugerido,
        filtro,
    )

    if not ruta:

        return False

    destino = Path(ruta)

    if destino.suffix.lower() not in (
        ".xlsx",
        ".csv",
    ):

        destino = destino.with_suffix(
            extension,
        )

    encabezados, filas = _filas_desde_registros(
        definition,
        registros,
    )

    try:

        if (
            destino.suffix.lower()
            == ".xlsx"
            and _TIENE_OPENPYXL
        ):

            _guardar_xlsx_datos(
                encabezados,
                filas,
                destino,
            )

        else:

            import csv

            if destino.suffix.lower() != ".csv":

                destino = destino.with_suffix(
                    ".csv",
                )

            with destino.open(
                "w",
                encoding="utf-8-sig",
                newline="",
            ) as archivo:

                escritor = csv.writer(
                    archivo,
                    delimiter=";",
                )

                escritor.writerow(
                    encabezados,
                )

                escritor.writerows(
                    filas,
                )

    except OSError as error:

        QMessageBox.critical(
            parent,
            "Exportar Excel",
            f"No se pudo guardar el archivo:\n{error}",
        )

        return False

    QMessageBox.information(
        parent,
        "Exportar Excel",
        f"Archivo exportado:\n{destino}",
    )

    return True


def _html_listado_registros(
    titulo: str,
    encabezados: list[str],
    filas: list[list[str]],
) -> str:

    celdas_encabezado = "".join(
        f"<th>{encabezado}</th>"
        for encabezado in encabezados
    )

    filas_html = []

    for fila in filas:

        celdas = "".join(
            f"<td>{valor}</td>"
            for valor in fila
        )

        filas_html.append(
            f"<tr>{celdas}</tr>",
        )

    return f"""
    <html>
    <head>
    <style>
        body {{ font-family: Arial, sans-serif; }}
        h2 {{ color: #1B4F8A; }}
        table {{
            border-collapse: collapse;
            width: 100%;
        }}
        th, td {{
            border: 1px solid #D1D5DB;
            padding: 6px;
            font-size: 10pt;
        }}
        th {{
            background: #1B4F8A;
            color: white;
        }}
        tr:nth-child(even) {{
            background: #F9FAFB;
        }}
    </style>
    </head>
    <body>
        <h2>{titulo}</h2>
        <table>
            <thead>
                <tr>{celdas_encabezado}</tr>
            </thead>
            <tbody>
                {"".join(filas_html)}
            </tbody>
        </table>
    </body>
    </html>
    """


def exportar_registros_pdf(
    definition,
    registros: list,
    *,
    parent: QWidget | None,
    titulo: str,
    nombre_archivo: str | None = None,
) -> bool:

    if not registros:

        QMessageBox.warning(
            parent,
            "Exportar PDF",
            "No hay datos para exportar.",
        )

        return False

    sugerido = (
        nombre_archivo
        or titulo.replace(" ", "_").lower()
    )

    if not sugerido.endswith(
        ".pdf",
    ):

        sugerido = f"{sugerido}.pdf"

    ruta, _ = QFileDialog.getSaveFileName(
        parent,
        "Exportar PDF",
        sugerido,
        "PDF (*.pdf)",
    )

    if not ruta:

        return False

    destino = Path(ruta)

    if destino.suffix.lower() != ".pdf":

        destino = destino.with_suffix(
            ".pdf",
        )

    encabezados, filas = _filas_desde_registros(
        definition,
        registros,
    )

    documento = QTextDocument()
    documento.setHtml(
        _html_listado_registros(
            titulo,
            encabezados,
            filas,
        ),
    )

    impresora = QPrinter(
        QPrinter.PrinterMode.HighResolution,
    )
    impresora.setOutputFormat(
        QPrinter.OutputFormat.PdfFormat,
    )
    impresora.setOutputFileName(
        str(
            destino,
        ),
    )
    impresora.setPageSize(
        QPageSize(
            QPageSize.PageSizeId.A4,
        ),
    )
    impresora.setPageMargins(
        QSizeF(
            12,
            12,
            12,
            12,
        ),
        QPrinter.Unit.Millimeter,
    )

    try:

        documento.print_(
            impresora,
        )

    except Exception as error:

        QMessageBox.critical(
            parent,
            "Exportar PDF",
            f"No se pudo generar el PDF:\n{error}",
        )

        return False

    if not destino.is_file():

        QMessageBox.warning(
            parent,
            "Exportar PDF",
            "No se pudo generar el archivo PDF.",
        )

        return False

    QMessageBox.information(
        parent,
        "Exportar PDF",
        f"Archivo exportado:\n{destino}",
    )

    return True


def exportar_registros_csv(
    definition,
    registros: list,
    *,
    parent: QWidget | None,
    titulo: str,
    nombre_archivo: str | None = None,
) -> bool:

    if not registros:

        QMessageBox.warning(
            parent,
            "Exportar CSV",
            "No hay datos para exportar.",
        )

        return False

    sugerido = (
        nombre_archivo
        or titulo.replace(" ", "_").lower()
    )

    if not sugerido.endswith(
        ".csv",
    ):

        sugerido = f"{sugerido}.csv"

    ruta, _ = QFileDialog.getSaveFileName(
        parent,
        "Exportar CSV",
        sugerido,
        "CSV (*.csv)",
    )

    if not ruta:

        return False

    destino = Path(ruta)

    if destino.suffix.lower() != ".csv":

        destino = destino.with_suffix(
            ".csv",
        )

    encabezados, filas = _filas_desde_registros(
        definition,
        registros,
    )

    try:

        import csv

        with destino.open(
            "w",
            encoding="utf-8-sig",
            newline="",
        ) as archivo:

            escritor = csv.writer(
                archivo,
                delimiter=";",
            )

            escritor.writerow(
                encabezados,
            )

            escritor.writerows(
                filas,
            )

    except OSError as error:

        QMessageBox.critical(
            parent,
            "Exportar CSV",
            f"No se pudo guardar el archivo:\n{error}",
        )

        return False

    QMessageBox.information(
        parent,
        "Exportar CSV",
        f"Archivo exportado:\n{destino}",
    )

    return True
